from flask import Flask, render_template, request, redirect, session, send_from_directory, jsonify
import psycopg2
import psycopg2.extras
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import os
import json
from ai_services import ai_validate_application_detailed, generate_hearing_schedule,   auto_assign_experts
from file_validator import validate_all_documents
import urllib.parse

app = Flask(__name__)



# Добавьте эту строку для статических файлов
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

app.secret_key = 'your_secret_key_here_change_in_production'

# ================= ИНТЕГРАЦИЯ С МОДУЛЕМ ПОИСКА =================
# Локально модуль поиска работает на 5000, модуль автоматизации — на 8000.
# Если запускаете модули на разных компьютерах, замените 127.0.0.1 на IP нужного компьютера.
SCOUTING_BASE_URL = os.getenv("SCOUTING_BASE_URL", "http://127.0.0.1:5000").rstrip("/")
AUTOMATION_BASE_URL = os.getenv("AUTOMATION_BASE_URL", "http://127.0.0.1:8000").rstrip("/")


# Конфигурация загрузки файлов
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ================= БАЗА ДАННЫХ =================
def get_db_connection():
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        return psycopg2.connect(database_url)
    else:
        # Для локальной разработки
        return psycopg2.connect(
            dbname="grants_db",
            user="postgres",
            password="1234",
            host="localhost",
            port="5432"
        )




# ================= INIT DB =================
def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Добавляем недостающие колонки если их нет
    columns_to_add = {
        'user_id': 'INTEGER',
        'score': 'INTEGER',
        'validation_errors': 'TEXT',
        'consent_file': 'TEXT',
        'education_file': 'TEXT'
    }

    for col, col_type in columns_to_add.items():
        cursor.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name='applications' AND column_name=%s
        """, (col,))
        if not cursor.fetchone():
            print(f"Добавляем столбец {col} в таблицу applications...")
            cursor.execute(f"ALTER TABLE applications ADD COLUMN {col} {col_type}")

    # Добавляем внешний ключ если его нет
    cursor.execute("""
        SELECT constraint_name 
        FROM information_schema.table_constraints 
        WHERE table_name='applications' AND constraint_type='FOREIGN KEY'
    """)
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE applications ADD CONSTRAINT fk_user FOREIGN KEY (user_id) REFERENCES users(id)")

    # Добавляем админа если его нет
    cursor.execute("SELECT * FROM users WHERE username = 'admin'")
    if not cursor.fetchone():
        cursor.execute("""
            INSERT INTO users (username, password, role, full_name)
            VALUES (%s, %s, 'admin', %s)
        """, ('admin', generate_password_hash('admin123'), 'Главный организатор'))

    conn.commit()
    cursor.close()
    conn.close()


init_db()


# ================= ДЕКОРАТОРЫ =================
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect("/login")
        return f(*args, **kwargs)

    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            return redirect("/login")
        return f(*args, **kwargs)

    return wrapper


# ================= ВАЛИДАЦИЯ =================
def validate_application(data):
    errors = []
    score = 100

    required_fields = ['full_name', 'project_name', 'amount', 'direction', 'summary', 'problem']
    for field in required_fields:
        if not data.get(field):
            errors.append(f"Отсутствует поле: {field}")
            score -= 15

    try:
        amount = float(data.get('amount', 0))
        if amount <= 0:
            errors.append("Сумма должна быть положительной")
            score -= 10
        elif amount > 10000000:
            errors.append("Сумма превышает максимальную (10 млн)")
            score -= 5
    except:
        errors.append("Некорректная сумма")
        score -= 10

    return {
        'errors': errors,
        'score': max(0, score)
    }


# ================= ОСНОВНЫЕ МАРШРУТЫ =================
@app.route("/")
def home():
    if session.get("user_id"):
        if session.get("role") == "admin":
            return redirect("/admin/dashboard")
        elif session.get("role") == "expert":
            return redirect("/expert/dashboard")
        else:
            # Для обычных пользователей - показываем главную страницу с конкурсами
            # НЕ делаем редирект на себя!
            return render_template("index.html")
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["full_name"] = user["full_name"]
            session["username"] = user["username"]

            if user["role"] == "admin":
                return redirect("/admin/dashboard")
            return redirect("/profile")

        return render_template("login.html", error="Неверный логин или пароль")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username")
        password = generate_password_hash(request.form.get("password"))
        full_name = request.form.get("full_name")
        email = request.form.get("email")

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO users (username, password, role, full_name, email)
                VALUES (%s, %s, 'applicant', %s, %s)
            """, (username, password, full_name, email))
            conn.commit()
            return redirect("/login")
        except psycopg2.IntegrityError:
            conn.rollback()
            return render_template("register.html", error="Пользователь уже существует")
        finally:
            cursor.close()
            conn.close()

    return render_template("register.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/apply/<int:contest_id>", methods=["GET", "POST"])
@login_required
def apply_with_contest(contest_id):
    """Подача заявки на конкретный конкурс"""
    if session.get("role") != "applicant":
        return redirect("/login")

    # Получаем информацию о конкурсе
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM contests WHERE id = %s", (contest_id,))
    contest = cursor.fetchone()
    cursor.close()

    if not contest:
        return "Конкурс не найден", 404

    if request.method == "POST":
        # Получаем данные из формы
        data = request.form.to_dict()

        # Обработка файлов
        consent_file = request.files.get('consent_file')
        education_file = request.files.get('education_file')

        consent_filename = None
        education_filename = None

        if consent_file and allowed_file(consent_file.filename):
            consent_filename = secure_filename(
                f"consent_{session['user_id']}_{datetime.now().timestamp()}_{consent_file.filename}")
            consent_file.save(os.path.join(app.config['UPLOAD_FOLDER'], consent_filename))

        if education_file and allowed_file(education_file.filename):
            education_filename = secure_filename(
                f"education_{session['user_id']}_{datetime.now().timestamp()}_{education_file.filename}")
            education_file.save(os.path.join(app.config['UPLOAD_FOLDER'], education_filename))

        # Валидация текстовых полей
        validation = validate_application(data)

        # AI проверка документов
        doc_validation_json = None
        if consent_filename and education_filename:
            try:
                consent_path = os.path.join(app.config['UPLOAD_FOLDER'], consent_filename)
                education_path = os.path.join(app.config['UPLOAD_FOLDER'], education_filename)

                applicant_check_data = {
                    'full_name': data.get('full_name'),
                    'passport_series': data.get('passport_series'),
                    'passport_number': data.get('passport_number'),
                    'education_place': data.get('education_place')
                }

                doc_validation_result = validate_all_documents(consent_path, education_path, applicant_check_data)

                if doc_validation_result['total_errors'] > 0:
                    for err in doc_validation_result['consent']['errors']:
                        validation['errors'].append(f"Согласие на ОПД: {err}")
                    for err in doc_validation_result['education']['errors']:
                        validation['errors'].append(f"Справка об обучении: {err}")
                    validation['score'] = max(0, validation['score'] - 20)

                doc_validation_json = json.dumps({
                    'consent': doc_validation_result['consent'],
                    'education': doc_validation_result['education'],
                    'overall_valid': doc_validation_result['overall_valid'],
                    'total_errors': doc_validation_result['total_errors'],
                    'total_warnings': doc_validation_result['total_warnings']
                }, ensure_ascii=False, default=str)

            except Exception as e:
                print(f"Ошибка при проверке документов: {e}")
                doc_validation_json = json.dumps({'error': str(e)})

        # Сохраняем заявку
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name='applications'
        """)
        existing_columns = [row[0] for row in cursor.fetchall()]

        insert_fields = []
        insert_values = []

        field_mapping = {
            'user_id': session["user_id"],
            'contest_id': contest_id,
            'full_name': data.get('full_name'),
            'address': data.get('address'),
            'birth_date': data.get('birth_date'),
            'passport_series': data.get('passport_series'),
            'passport_number': data.get('passport_number'),
            'passport_issued_by': data.get('passport_issued_by'),
            'passport_date': data.get('passport_date'),
            'education_place': data.get('education_place'),
            'project_name': data.get('project_name'),
            'project_short': data.get('project_short'),
            'amount': data.get('amount'),
            'direction': data.get('direction'),
            'duration': data.get('duration'),
            'team_size': data.get('team_size'),
            'prototype': data.get('prototype'),
            'summary': data.get('summary'),
            'problem': data.get('problem'),
            'audience': data.get('audience'),
            'uniqueness': data.get('uniqueness'),
            'plan': data.get('plan'),
            'results': data.get('results'),
            'score': validation['score'],
            'validation_errors': ', '.join(validation['errors']) if validation['errors'] else None,
            'consent_file': consent_filename,
            'education_file': education_filename,
            'document_validation': doc_validation_json
        }

        for field, value in field_mapping.items():
            if field in existing_columns:
                insert_fields.append(field)
                insert_values.append(value)

        if 'status' in existing_columns:
            insert_fields.append('status')
            insert_values.append('На рассмотрении')

        if 'created_at' in existing_columns:
            insert_fields.append('created_at')
            insert_values.append(datetime.now())

        query = f"""
            INSERT INTO applications ({', '.join(insert_fields)})
            VALUES ({', '.join(['%s'] * len(insert_values))})
        """

        cursor.execute(query, insert_values)
        conn.commit()
        cursor.close()
        conn.close()

        return redirect("/profile")

    # Определяем тип формы для отображения
    form_type = 'simple'
    if contest_id == 2:  # Студенческие гранты - многостраничная форма
        form_type = 'multistep'

    return render_template(f"apply_{form_type}.html", contest=contest)

# ===== ПОДАЧА ЗАЯВКИ =====
@app.route("/apply", methods=["GET", "POST"])
@login_required
def apply():
    """Подача заявки (по умолчанию - общий конкурс)"""
    if session.get("role") != "applicant":
        return redirect("/login")

    # Если нет параметра конкурса, используем общий
    return apply_with_contest(1)  # или redirect с параметром


# ===== ПРОФИЛЬ ЗАЯВИТЕЛЯ (СПИСОК ЗАЯВОК) =====
@app.route("/profile")
@login_required
def profile():
    if session.get("role") != "applicant":
        return redirect("/login")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем заявки с названием конкурса
    cursor.execute("""
        SELECT a.id, a.project_name, a.amount, a.status, a.score, a.created_at, 
               a.contest_id, c.name as contest_name
        FROM applications a
        LEFT JOIN contests c ON a.contest_id = c.id
        WHERE a.user_id = %s 
        ORDER BY a.created_at DESC
    """, (session["user_id"],))

    apps = cursor.fetchall()

    # Получаем данные пользователя (email из базы)
    cursor.execute("""
        SELECT full_name, username, email 
        FROM users 
        WHERE id = %s
    """, (session["user_id"],))

    user_data = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template("profile.html",
                           applications=apps,
                           user=user_data)


# ================= АДМИН-ПАНЕЛЬ =================
@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    """Панель организатора с выбором конкурса"""
    contest_id = request.args.get('contest_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем список конкурсов
    cursor.execute("SELECT id, name FROM contests ORDER BY id")
    contests = cursor.fetchall()

    # Если конкурс не выбран, берем первый
    if not contest_id and contests:
        contest_id = contests[0]['id']

    # Получаем заявки ТОЛЬКО по выбранному конкурсу
    if contest_id:
        cursor.execute("""
            SELECT a.*, u.username as applicant_username 
            FROM applications a
            LEFT JOIN users u ON a.user_id = u.id
            WHERE a.contest_id = %s
            ORDER BY a.id DESC
        """, (contest_id,))
    else:
        # Если нет конкурсов вообще
        applications = []

    apps = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("admin_dashboard.html",
                           applications=apps,
                           contests=contests,
                           selected_contest=contest_id)


@app.route("/admin/application/<int:app_id>")
@admin_required
def admin_application(app_id):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT a.*, u.username as applicant_username, u.email as applicant_email
        FROM applications a
        LEFT JOIN users u ON a.user_id = u.id
        WHERE a.id = %s
    """, (app_id,))

    app_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if not app_data:
        return "Заявка не найдена", 404

    return render_template("admin_application_detail.html", app=app_data)


@app.route("/admin/update_status/<int:app_id>", methods=["POST"])
@admin_required
def update_status(app_id):
    new_status = request.form.get('status')

    if new_status not in ['Рассмотрение', 'Одобрено', 'Отклонено']:
        return "Неверный статус", 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE applications SET status = %s WHERE id = %s", (new_status, app_id))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(f"/admin/application/{app_id}")


# ================= ЭКСПОРТ ОТЧЕТОВ =================

@app.route("/admin/export_report/<format>")
@admin_required
def export_report(format):
    """Экспорт отчета в PDF или Excel"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем все заявки с сортировкой по баллам
    cursor.execute("""
        SELECT id, project_name, full_name, amount, direction, score, status, created_at
        FROM applications
        ORDER BY score DESC NULLS LAST
    """)
    applications = cursor.fetchall()
    cursor.close()
    conn.close()

    if format == 'excel':
        return export_to_excel(applications)
    elif format == 'pdf':
        return export_to_pdf(applications)
    else:
        return "Неверный формат", 400


def export_to_excel(applications):
    """Экспорт в Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по заявкам"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ['ID', 'Проект', 'Заявитель', 'Сумма (руб)', 'Направление', 'Балл', 'Статус', 'Дата подачи']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app['id']).border = thin_border
        ws.cell(row=row, column=2, value=app['project_name']).border = thin_border
        ws.cell(row=row, column=3, value=app['full_name']).border = thin_border
        ws.cell(row=row, column=4, value=float(app['amount']) if app['amount'] else 0).border = thin_border
        ws.cell(row=row, column=5, value=app['direction']).border = thin_border
        ws.cell(row=row, column=6, value=app['score'] if app['score'] else 0).border = thin_border

        # Цвет для статуса
        status_cell = ws.cell(row=row, column=7, value=app['status'])
        status_cell.border = thin_border
        if app['status'] == 'Одобрено':
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        elif app['status'] == 'Отклонено':
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")

        created_at = app['created_at'].strftime('%d.%m.%Y') if app['created_at'] else ''
        ws.cell(row=row, column=8, value=created_at).border = thin_border

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Добавляем статистику на отдельный лист
    ws_stats = wb.create_sheet("Статистика")

    # Подсчет статистики
    total = len(applications)
    approved = sum(1 for app in applications if app['status'] == 'Одобрено')
    rejected = sum(1 for app in applications if app['status'] == 'Отклонено')
    review = sum(1 for app in applications if app['status'] == 'На рассмотрении')

    scores = [app['score'] for app in applications if app['score'] is not None]
    avg_score = sum(scores) / len(scores) if scores else 0
    max_score = max(scores) if scores else 0

    stats_data = [
        ['Показатель', 'Значение'],
        ['Всего заявок', total],
        ['Одобрено', approved],
        ['Отклонено', rejected],
        ['На рассмотрении', review],
        ['Средний балл', f'{avg_score:.2f}'],
        ['Максимальный балл', max_score]
    ]

    for row, data in enumerate(stats_data, 1):
        for col, value in enumerate(data, 1):
            cell = ws_stats.cell(row=row, column=col, value=value)
            if row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")

    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20

    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=report.xlsx'}
    )


def export_to_pdf(applications):
    """Экспорт в PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    # Регистрируем русский шрифт
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
        font_name = 'DejaVu'
    except:
        font_name = 'Helvetica'

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()

    # Стили
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=20,
        fontName=font_name
    )

    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        fontName=font_name
    )

    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=9,
        fontName=font_name,
        textColor=colors.white
    )

    story = []

    # Заголовок
    story.append(Paragraph("Отчет по заявкам на грантовый конкурс", title_style))
    story.append(Paragraph(f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 20))

    # Статистика
    total = len(applications)
    approved = sum(1 for app in applications if app['status'] == 'Одобрено')
    rejected = sum(1 for app in applications if app['status'] == 'Отклонено')
    review = sum(1 for app in applications if app['status'] == 'На рассмотрении')

    scores = [app['score'] for app in applications if app['score'] is not None]
    avg_score = sum(scores) / len(scores) if scores else 0

    stats_data = [
        ['Всего заявок', str(total)],
        ['Одобрено', str(approved)],
        ['Отклонено', str(rejected)],
        ['На рассмотрении', str(review)],
        ['Средний балл', f'{avg_score:.2f}']
    ]

    stats_table = Table(stats_data, colWidths=[5 * cm, 5 * cm])
    stats_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 20))

    # Таблица заявок
    data = [['ID', 'Проект', 'Заявитель', 'Сумма', 'Балл', 'Статус']]

    for app in applications[:50]:  # Ограничиваем 50 заявками для PDF
        status_text = app['status']
        status_color = colors.green if status_text == 'Одобрено' else (
            colors.red if status_text == 'Отклонено' else colors.orange)

        data.append([
            str(app['id']),
            app['project_name'][:40] if app['project_name'] else '',
            app['full_name'][:30] if app['full_name'] else '',
            f"{float(app['amount']):,.0f}" if app['amount'] else '0',
            str(app['score']) if app['score'] else '0',
            status_text
        ])

    table = Table(data, colWidths=[1.5 * cm, 5 * cm, 5 * cm, 3 * cm, 1.5 * cm, 3 * cm])

    table_style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    table.setStyle(table_style)
    story.append(table)

    # Постраничное отображение победителей для большого количества
    winners = [app for app in applications if app['status'] == 'Одобрено'][:5]
    if winners:
        story.append(PageBreak())
        story.append(Paragraph("Рекомендованные к финансированию проекты", title_style))
        story.append(Spacer(1, 10))

        winner_data = [['Место', 'Проект', 'Заявитель', 'Балл']]
        for idx, app in enumerate(winners, 1):
            winner_data.append([
                str(idx),
                app['project_name'][:50] if app['project_name'] else '',
                app['full_name'][:30] if app['full_name'] else '',
                str(app['score']) if app['score'] else '0'
            ])

        winner_table = Table(winner_data, colWidths=[2 * cm, 7 * cm, 5 * cm, 2 * cm])
        winner_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(winner_table)

    doc.build(story)
    output.seek(0)

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=report.pdf'}
    )


@app.route("/admin/export_winners")
@admin_required
def export_winners():
    """Экспорт списка победителей"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT id, project_name, full_name, amount, score, created_at
        FROM applications
        WHERE status = 'Одобрено'
        ORDER BY score DESC
        LIMIT 10
    """)
    winners = cursor.fetchall()
    cursor.close()
    conn.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Победители грантового конкурса"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")

    headers = ['Место', 'ID', 'Проект', 'Заявитель', 'Сумма (руб)', 'Балл', 'Дата подачи']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, winner in enumerate(winners, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=winner['id'])
        ws.cell(row=row, column=3, value=winner['project_name'])
        ws.cell(row=row, column=4, value=winner['full_name'])
        ws.cell(row=row, column=5, value=float(winner['amount']) if winner['amount'] else 0)
        ws.cell(row=row, column=6, value=winner['score'] if winner['score'] else 0)
        created_at = winner['created_at'].strftime('%d.%m.%Y') if winner['created_at'] else ''
        ws.cell(row=row, column=7, value=created_at)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=winners.xlsx'}
    )


@app.route("/admin/download/<filename>")
@admin_required
def download_file(filename):
    """Скачивание загруженных файлов"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# ================= НОВЫЕ AI МАРШРУТЫ =================

@app.route("/admin/ai_check_all")
@admin_required
def ai_check_all():
    """AI проверка заявок по конкретному конкурсу"""
    contest_id = request.args.get('contest_id', type=int)

    if not contest_id:
        return "Не выбран конкурс", 400

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Проверяем ТОЛЬКО заявки выбранного конкурса
    cursor.execute("""
        SELECT * FROM applications 
        WHERE status = 'На рассмотрении' AND contest_id = %s
    """, (contest_id,))

    applications = cursor.fetchall()

    results = []
    for app in applications:
        # Детальная AI проверка
        ai_result = ai_validate_application_detailed(dict(app))

        # Автоматическое обновление статуса на основе AI
        new_status = 'На рассмотрении'
        if ai_result['can_approve'] and ai_result['total_score'] >= 75:
            new_status = 'Одобрено'
        elif ai_result['total_score'] < 50:
            new_status = 'Отклонено'

        # Сохраняем полный результат проверки (СОХРАНЯЕМ formal_score)
        cursor.execute("""
            UPDATE applications 
            SET score = %s, 
                formal_score = %s,
                validation_errors = %s,
                status = %s
            WHERE id = %s
        """, (
            ai_result['total_score'],
            ai_result['total_score'],  # formal_score = балл AI проверки
            json.dumps(ai_result, ensure_ascii=False, default=str),
            new_status,
            app['id']
        ))

        results.append({
            'id': app['id'],
            'project_name': app['project_name'],
            'full_name': app['full_name'],
            'score': ai_result['total_score'],
            'new_status': new_status,
            'summary': ai_result['summary'],
            'recommendations': ai_result['recommendations'][:3],
            'old_status': app['status']
        })

    conn.commit()
    cursor.close()
    conn.close()

    return render_template("ai_check_report.html", results=results, total=len(results))

# ================= НАЗНАЧЕНИЕ ЭКСПЕРТОВ =================

@app.route("/admin/experts")
@admin_required
def experts_list():
    """Список экспертов"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("SELECT * FROM experts ORDER BY full_name")
    experts = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("experts_list.html", experts=experts)


@app.route("/admin/experts/add", methods=["GET", "POST"])
@admin_required
def add_expert():
    """Добавление эксперта"""
    if request.method == "POST":
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        expertise_areas = request.form.getlist('expertise_areas')
        max_load = request.form.get('max_load', 5)
        rating = request.form.get('rating', 5.0)

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO experts (full_name, email, expertise_areas, max_load, rating, is_available)
            VALUES (%s, %s, %s, %s, %s, TRUE)
        """, (full_name, email, expertise_areas, max_load, rating))

        conn.commit()
        cursor.close()
        conn.close()

        return redirect("/admin/experts")

    return render_template("add_expert.html")


@app.route("/admin/assign_experts")
@admin_required
def assign_experts():
    """Автоматическое назначение экспертов на заявки"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем заявки, требующие назначения экспертов
    cursor.execute("""
        SELECT a.* 
        FROM applications a
        LEFT JOIN expert_assignments ea ON a.id = ea.application_id
        WHERE ea.id IS NULL AND a.status = 'Одобрено'
        ORDER BY a.score DESC
    """)
    pending_apps = cursor.fetchall()

    # Получаем всех экспертов
    cursor.execute("SELECT * FROM experts WHERE is_available = TRUE ORDER BY current_load ASC")
    experts = cursor.fetchall()

    if not experts:
        cursor.close()
        conn.close()
        return "Нет доступных экспертов. Добавьте экспертов в систему.", 400

    assignments_results = []

    for app in pending_apps:
        # Используем AI для назначения экспертов
        assigned_experts = auto_assign_experts(dict(app), [dict(e) for e in experts])

        for expert_info in assigned_experts:
            # Обновляем нагрузку эксперта
            cursor.execute("""
                UPDATE experts 
                SET current_load = current_load + 1 
                WHERE id = %s
            """, (expert_info['expert_id'],))

            # Создаем назначение
            cursor.execute("""
                INSERT INTO expert_assignments (application_id, expert_id, status, evaluation_deadline)
                VALUES (%s, %s, 'assigned', NOW() + INTERVAL '14 days')
                RETURNING id
            """, (app['id'], expert_info['expert_id']))

            assignment_id = cursor.fetchone()[0]

            assignments_results.append({
                'application_id': app['id'],
                'project_name': app['project_name'],
                'expert_name': expert_info['full_name'],
                'score': expert_info['score'],
                'assignment_id': assignment_id
            })

    conn.commit()
    cursor.close()
    conn.close()

    return render_template("assign_experts_result.html",
                           assignments=assignments_results,
                           total_apps=len(pending_apps))


@app.route("/admin/generate_schedule")
@admin_required
def generate_schedule():
    """Формирование графика заслушиваний"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT id, project_name, full_name 
        FROM applications 
        WHERE status = 'Одобрено' AND score >= 70
        ORDER BY score DESC
    """)
    approved_apps = cursor.fetchall()

    schedule = generate_hearing_schedule(approved_apps)
    cursor.close()
    conn.close()

    # Простое отображение
    html = "<h1>График заслушиваний</h1><table border='1' cellpadding='10'>"
    html += "<tr><th>Дата</th><th>Время</th><th>Проект</th><th>Заявитель</th></tr>"
    for item in schedule:
        html += f"<tr><td>{item['date']}</td><td>{item['time']}</td><td>{item['project_name']}</td><td>{item['applicant']}</td></tr>"
    html += "</table><br><a href='/admin/dashboard'>← Назад</a>"

    return html


# ================= НАЗНАЧЕНИЕ ЭКСПЕРТОВ (НОВАЯ ВЕРСИЯ) =================

@app.route("/admin/assign_experts_batch")
@admin_required
def assign_experts_batch():
    """Массовое назначение экспертов на проекты (один проект - несколько экспертов)"""
    contest_id = request.args.get('contest_id')

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем одобренные проекты без назначенных экспертов
    if contest_id:
        cursor.execute("""
            SELECT a.* 
            FROM applications a
            LEFT JOIN project_experts pe ON a.id = pe.application_id
            WHERE pe.id IS NULL AND a.status = 'Одобрено' AND a.contest_id = %s
            ORDER BY a.score DESC
        """, (contest_id,))
    else:
        cursor.execute("""
            SELECT a.* 
            FROM applications a
            LEFT JOIN project_experts pe ON a.id = pe.application_id
            WHERE pe.id IS NULL AND a.status = 'Одобрено'
            ORDER BY a.score DESC
        """)

    pending_apps = cursor.fetchall()

    # Получаем всех доступных экспертов
    cursor.execute("SELECT * FROM experts WHERE is_available = TRUE ORDER BY rating DESC")
    experts = cursor.fetchall()

    if not experts:
        cursor.close()
        conn.close()
        return "Нет доступных экспертов. Добавьте экспертов в систему.", 400

    results = []

    for app in pending_apps:
        # Используем AI для назначения экспертов
        assigned_experts = auto_assign_experts(dict(app), [dict(e) for e in experts])

        expert_ids = [e['expert_id'] for e in assigned_experts]
        expert_names = [e['full_name'] for e in assigned_experts]

        # Обновляем нагрузку экспертов
        for expert_info in assigned_experts:
            cursor.execute("""
                UPDATE experts 
                SET current_load = current_load + 1 
                WHERE id = %s
            """, (expert_info['expert_id'],))

        # Создаем запись о назначении
        cursor.execute("""
            INSERT INTO project_experts (application_id, expert_ids, status, contest_id)
            VALUES (%s, %s, 'pending', %s)
            RETURNING id
        """, (app['id'], expert_ids, contest_id))

        results.append({
            'application_id': app['id'],
            'project_name': app['project_name'],
            'expert_names': expert_names,
            'direction': app['direction']
        })

    conn.commit()
    cursor.close()
    conn.close()

    return render_template("assign_experts_batch.html", assignments=results, total=len(results))

@app.route("/admin/project_experts")
@admin_required
def project_experts_list():
    """Список проектов с назначенными экспертами"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT 
            pe.id,
            pe.application_id,
            pe.expert_ids,
            pe.status,
            pe.created_at,
            a.project_name,
            a.full_name as applicant_name,
            a.direction,
            a.score as ai_score
        FROM project_experts pe
        JOIN applications a ON pe.application_id = a.id
        ORDER BY pe.created_at DESC
    """)

    rows = cursor.fetchall()
    assignments = []

    for row in rows:
        assignment = dict(row)
        expert_ids = assignment.get('expert_ids', [])

        # Получаем данные экспертов
        if expert_ids and len(expert_ids) > 0:
            cursor.execute("SELECT id, full_name, email FROM experts WHERE id = ANY(%s)", (expert_ids,))
            experts_data = cursor.fetchall()
            assignment['experts_detail'] = [dict(e) for e in experts_data]
        else:
            assignment['experts_detail'] = []

        assignments.append(assignment)

    cursor.close()
    conn.close()

    return render_template("project_experts_list.html", assignments=assignments)

# ================= ЭКСПЕРТНЫЙ КАБИНЕТ =================
# ===== ДОБАВИТЬ РЕГИСТРАЦИЮ ЭКСПЕРТОВ =====
@app.route("/expert/register", methods=["GET", "POST"])
def expert_register():
    """Регистрация нового эксперта"""
    if request.method == "POST":
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        password = generate_password_hash(request.form.get('password'))
        phone = request.form.get('phone')
        position = request.form.get('position')
        organization = request.form.get('organization')
        experience_years = request.form.get('experience_years', 0)
        education = request.form.get('education')
        achievements = request.form.get('achievements')
        publications = request.form.get('publications')
        expertise_areas = request.form.getlist('expertise_areas')

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO experts (full_name, email, password, phone, position, organization, 
                                     experience_years, education, achievements, publications, 
                                     expertise_areas, is_available, rating, current_load, max_load)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, 5.0, 0, 5)
                RETURNING id
            """, (full_name, email, password, phone, position, organization,
                  experience_years, education, achievements, publications, expertise_areas))

            expert_id = cursor.fetchone()[0]
            conn.commit()

            # Автоматический вход после регистрации
            session['expert_id'] = expert_id
            session['expert_name'] = full_name
            session['role'] = 'expert'

            return redirect("/expert/dashboard")

        except psycopg2.IntegrityError:
            conn.rollback()
            return render_template("expert_register.html", error="Эксперт с таким email уже существует")
        finally:
            cursor.close()
            conn.close()

    # GET - показываем форму регистрации
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT id, name FROM contests")
    contests = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("expert_register.html", contests=contests)


@app.route("/expert/login", methods=["GET", "POST"])
def expert_login():
    """Вход для экспертов (по email и паролю)"""
    if request.method == "POST":
        email = request.form.get('email')
        password = request.form.get('password')

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT * FROM experts WHERE email = %s", (email,))
        expert = cursor.fetchone()
        cursor.close()
        conn.close()

        if expert and check_password_hash(expert['password'], password):
            session['expert_id'] = expert['id']
            session['expert_name'] = expert['full_name']
            session['expert_email'] = expert['email']
            session['role'] = 'expert'
            return redirect("/expert/dashboard")

        return render_template("expert_login.html", error="Неверный email или пароль")

    return render_template("expert_login.html")


# ===== ДОБАВИТЬ МАРШРУТ ДЛЯ ПРОФИЛЯ ЭКСПЕРТА =====
@app.route("/expert/profile", methods=["GET", "POST"])
def expert_profile():
    """Профиль эксперта с возможностью редактирования"""
    if not session.get('expert_id'):
        return redirect("/expert/login")

    expert_id = session['expert_id']
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == "POST":
        # Обновление профиля
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        position = request.form.get('position')
        organization = request.form.get('organization')
        experience_years = request.form.get('experience_years')
        education = request.form.get('education')
        achievements = request.form.get('achievements')
        publications = request.form.get('publications')
        expertise_areas = request.form.getlist('expertise_areas')

        cursor.execute("""
            UPDATE experts 
            SET full_name = %s, phone = %s, position = %s, organization = %s,
                experience_years = %s, education = %s, achievements = %s, 
                publications = %s, expertise_areas = %s
            WHERE id = %s
        """, (full_name, phone, position, organization, experience_years,
              education, achievements, publications, expertise_areas, expert_id))
        conn.commit()

        session['expert_name'] = full_name

    # GET - показываем профиль
    cursor.execute("SELECT * FROM experts WHERE id = %s", (expert_id,))
    expert = cursor.fetchone()

    cursor.execute("SELECT id, name FROM contests")
    contests = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("expert_profile.html", expert=expert, contests=contests)


@app.route("/expert/dashboard")
def expert_dashboard():
    if not session.get('expert_id'):
        return redirect("/expert/login")

    expert_id = session['expert_id']
    expert_email = session.get('expert_email', '')
    expert_name = session.get('expert_name', '')

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем проекты для оценки
    cursor.execute("""
        SELECT pe.id as project_expert_id, a.project_name, a.full_name as applicant_name, 
               a.direction, pe.status,
               CASE WHEN EXISTS (
                   SELECT 1 FROM expert_scores es 
                   WHERE es.project_expert_id = pe.id AND es.expert_id = %s
               ) THEN TRUE ELSE FALSE END as has_scores
        FROM project_experts pe
        JOIN applications a ON pe.application_id = a.id
        WHERE %s = ANY(pe.expert_ids)
        ORDER BY pe.created_at DESC
    """, (expert_id, expert_id))

    projects = cursor.fetchall()

    # Получаем уведомления
    cursor.execute("""
        SELECT en.*, a.project_name
        FROM expert_notifications en
        JOIN project_experts pe ON en.project_expert_id = pe.id
        JOIN applications a ON pe.application_id = a.id
        WHERE en.expert_id = %s AND en.is_resolved = FALSE
    """, (expert_id,))

    notifications = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("expert_dashboard.html",
                           projects=projects,
                           notifications=notifications,
                           expert_name=expert_name,
                           expert_email=expert_email)


@app.route("/expert/evaluate/<int:project_expert_id>", methods=["GET", "POST"])
def expert_evaluate(project_expert_id):
    """Страница оценки проекта экспертом"""
    if not session.get('expert_id'):
        return redirect("/expert/login")

    expert_id = session['expert_id']

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем информацию о проекте
    cursor.execute("""
        SELECT 
            pe.*,
            a.project_name,
            a.full_name as applicant_name,
            a.summary,
            a.problem,
            a.uniqueness,
            a.plan,
            a.results,
            a.direction,
            a.amount,
            a.team_size
        FROM project_experts pe
        JOIN applications a ON pe.application_id = a.id
        WHERE pe.id = %s
    """, (project_expert_id,))

    project = cursor.fetchone()

    if not project or expert_id not in project['expert_ids']:
        return "Доступ запрещен", 403

    if request.method == "POST":
        # Сохраняем оценки по критериям
        criteria = ['Актуальность', 'Новизна', 'Реализуемость', 'Команда', 'Бюджет']

        for criterion in criteria:
            score = request.form.get(f'score_{criterion}')
            comment = request.form.get(f'comment_{criterion}')

            if score:
                cursor.execute("""
                    INSERT INTO expert_scores (project_expert_id, expert_id, criteria_name, score, comment)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (project_expert_id, expert_id, criteria_name) 
                    DO UPDATE SET score = EXCLUDED.score, comment = EXCLUDED.comment
                """, (project_expert_id, expert_id, criterion, int(score), comment))

        # Обновляем статус проекта в project_experts
        cursor.execute("""
            UPDATE project_experts 
            SET status = 'in_progress' 
            WHERE id = %s
        """, (project_expert_id,))

        # Закрываем уведомления для этого эксперта по этому проекту
        cursor.execute("""
            UPDATE expert_notifications 
            SET is_resolved = TRUE, resolved_at = NOW()
            WHERE expert_id = %s AND project_expert_id = %s AND is_resolved = FALSE
        """, (expert_id, project_expert_id))

        conn.commit()

        # Проверяем, все ли эксперты оценили проект
        cursor.execute("""
            SELECT COUNT(DISTINCT expert_id) as evaluated_count
            FROM expert_scores
            WHERE project_expert_id = %s
        """, (project_expert_id,))

        evaluated = cursor.fetchone()['evaluated_count']
        total_experts = len(project['expert_ids'])

        if evaluated >= total_experts:
            cursor.execute("""
                UPDATE project_experts 
                SET status = 'completed' 
                WHERE id = %s
            """, (project_expert_id,))
            conn.commit()

            # Запускаем анализ аномалий и формирование итогов
            check_and_finalize_project(project_expert_id, conn)

        conn.close()
        return redirect("/expert/dashboard")

    # ПОЛУЧАЕМ СОХРАНЕННЫЕ ОЦЕНКИ (ТОЛЬКО ДЛЯ GET-ЗАПРОСА)
    cursor.execute("""
        SELECT criteria_name, score, comment
        FROM expert_scores
        WHERE project_expert_id = %s AND expert_id = %s
    """, (project_expert_id, expert_id))

    saved_scores = {}
    for row in cursor.fetchall():
        saved_scores[row['criteria_name']] = {'score': row['score'], 'comment': row['comment']}

    cursor.close()
    conn.close()

    return render_template("expert_evaluate.html", project=project, saved_scores=saved_scores)


def check_and_finalize_project(project_expert_id, conn):
    """Проверка аномалий и формирование итогового результата"""
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем все оценки по проекту
    cursor.execute("""
        SELECT 
            es.*,
            e.full_name as expert_name,
            e.id as expert_id
        FROM expert_scores es
        JOIN experts e ON es.expert_id = e.id
        WHERE es.project_expert_id = %s
    """, (project_expert_id,))

    all_scores = cursor.fetchall()

    # Получаем информацию о проекте
    cursor.execute("""
        SELECT pe.application_id, a.project_name, a.formal_score, a.score as old_score, pe.expert_ids
        FROM project_experts pe
        JOIN applications a ON pe.application_id = a.id
        WHERE pe.id = %s
    """, (project_expert_id,))
    project_info = cursor.fetchone()

    # Анализ аномалий
    anomalies = []
    criteria_scores = {}

    for score in all_scores:
        criteria = score['criteria_name']
        if criteria not in criteria_scores:
            criteria_scores[criteria] = []
        criteria_scores[criteria].append({
            'score': score['score'],
            'expert_id': score['expert_id'],
            'expert_name': score['expert_name']
        })

    # Выявляем аномалии
    anomaly_detected = False
    anomaly_details = []

    for criteria, scores in criteria_scores.items():
        if len(scores) >= 2:
            score_values = [s['score'] for s in scores]
            mean = sum(score_values) / len(score_values)

            for s in scores:
                if abs(s['score'] - mean) > 3:
                    anomaly_detected = True
                    anomaly_details.append({
                        'criteria': criteria,
                        'expert_id': s['expert_id'],
                        'expert_name': s['expert_name'],
                        'score': s['score'],
                        'mean': round(mean, 1),
                        'deviation': round(abs(s['score'] - mean), 1)
                    })

                    cursor.execute("""
                        INSERT INTO expert_notifications (expert_id, project_expert_id, criteria, current_score, recommended_score, created_at)
                        VALUES (%s, %s, %s, %s, %s, NOW())
                    """, (s['expert_id'], project_expert_id, criteria, s['score'], round(mean, 1)))

    # Вычисляем балл экспертов
    all_values = [s['score'] for s in all_scores]
    if all_values:
        sorted_values = sorted(all_values)
        n = len(sorted_values)
        if n % 2 == 0:
            expert_avg = (sorted_values[n // 2 - 1] + sorted_values[n // 2]) / 2
        else:
            expert_avg = sorted_values[n // 2]
        expert_score = round(expert_avg * 10, 2)
    else:
        expert_score = 0

    final_score = expert_score

    # Определяем статус
    if anomaly_detected:
        final_status = 'На доработку'
        project_status = 'revision_needed'
    elif final_score >= 85:
        final_status = 'Победитель'
        project_status = 'completed'
    elif final_score >= 70:
        final_status = 'Рекомендовано'
        project_status = 'completed'
    else:
        final_status = 'Отклонено'
        project_status = 'rejected'

    # Сохраняем в project_ratings
    cursor.execute("""
        SELECT id FROM project_ratings 
        WHERE application_id = %s
    """, (project_info['application_id'],))

    existing = cursor.fetchone()

    report_data = json.dumps({
        'anomalies': anomaly_details,
        'criteria_scores': {k: [s['score'] for s in v] for k, v in criteria_scores.items()}
    })

    if existing:
        cursor.execute("""
            UPDATE project_ratings 
            SET average_score = %s, 
                final_status = %s,
                anomaly_detected = %s,
                report_data = %s,
                expert_score = %s
            WHERE application_id = %s
        """, (final_score, final_status, anomaly_detected, report_data, expert_score, project_info['application_id']))
    else:
        cursor.execute("""
            INSERT INTO project_ratings (application_id, average_score, final_status, anomaly_detected, report_data, expert_score)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (project_info['application_id'], final_score, final_status, anomaly_detected, report_data, expert_score))

    # Обновляем статус в project_experts
    cursor.execute("""
        UPDATE project_experts 
        SET status = %s
        WHERE id = %s
    """, (project_status, project_expert_id))

    # Обновляем заявку (НЕ ТРОГАЕМ formal_score)
    cursor.execute("""
        UPDATE applications 
        SET status = %s, 
            score = %s,
            expert_score = %s
        WHERE id = %s
    """, (final_status, final_score, expert_score, project_info['application_id']))

    conn.commit()
    cursor.close()

@app.route("/admin/anomalies_report")
@admin_required
def anomalies_report():
    """Отчет по аномалиям в оценках"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT 
            pr.*,
            a.project_name,
            a.full_name as applicant_name
        FROM project_ratings pr
        JOIN applications a ON pr.application_id = a.id
        WHERE pr.anomaly_detected = TRUE
        ORDER BY pr.created_at DESC
    """)

    anomalies = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("anomalies_report.html", anomalies=anomalies)


@app.route("/admin/project_final_report/<int:application_id>")
@admin_required
def project_final_report(application_id):
    """Итоговый отчет по проекту"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT pr.*, a.*
        FROM project_ratings pr
        JOIN applications a ON pr.application_id = a.id
        WHERE pr.application_id = %s
    """, (application_id,))

    report = cursor.fetchone()

    # Получаем оценки экспертов
    cursor.execute("""
        SELECT 
            e.full_name as expert_name,
            es.criteria_name,
            es.score
        FROM expert_scores es
        JOIN experts e ON es.expert_id = e.id
        WHERE es.project_expert_id IN (
            SELECT id FROM project_experts WHERE application_id = %s
        )
        ORDER BY e.full_name, es.criteria_name
    """, (application_id,))

    rows = cursor.fetchall()

    print(f"DEBUG: Найдено оценок: {len(rows)}")  # Отладка

    # Группируем оценки по экспертам
    expert_scores = {}
    for row in rows:
        expert = row['expert_name']
        if expert not in expert_scores:
            expert_scores[expert] = {}
        expert_scores[expert][row['criteria_name']] = row['score']

    # Вычисляем средние
    criteria_list = ['Актуальность', 'Новизна', 'Реализуемость', 'Команда', 'Бюджет']
    expert_averages = []
    criteria_sums = {c: 0 for c in criteria_list}
    expert_count = len(expert_scores)

    for expert, scores in expert_scores.items():
        expert_sum = 0
        for criteria in criteria_list:
            score = scores.get(criteria, 0)
            expert_sum += score
            criteria_sums[criteria] += score
        expert_avg = round(expert_sum / 5, 1) if expert_sum > 0 else 0
        expert_averages.append(expert_avg)

    total_avg = round(sum(expert_averages) / expert_count, 1) if expert_count > 0 else 0
    final_score_100 = round(total_avg * 10, 1)

    cursor.close()
    conn.close()

    return render_template("project_final_report.html",
                           report=report,
                           expert_scores=expert_scores,
                           total_avg=total_avg,
                           final_score_100=final_score_100)

@app.route("/expert/logout")
def expert_logout():
    session.pop('expert_id', None)
    session.pop('expert_name', None)
    session.pop('expert_email', None)
    session.pop('role', None)
    return redirect("/expert/login")


# ================= ИТОГОВЫЙ ОТЧЕТ ПОСЛЕ ЗАСЛУШИВАНИЯ =================

@app.route("/admin/final_hearing_report")
@admin_required
def final_hearing_report():
    """Итоговый отчет после заслушивания со всеми экспертными оценками"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    contest_id = request.args.get('contest_id')

    # Получаем все проекты, у которых есть экспертные оценки
    if contest_id:
        cursor.execute("""
            SELECT DISTINCT 
                a.id as application_id,
                a.project_name,
                a.full_name as applicant_name,
                a.direction,
                pr.average_score as final_score,
                pr.final_status,
                pr.report_data,
                pr.anomaly_detected
            FROM applications a
            LEFT JOIN project_ratings pr ON a.id = pr.application_id
            LEFT JOIN project_experts pe ON a.id = pe.application_id
            WHERE (pe.id IS NOT NULL OR pr.id IS NOT NULL) AND a.contest_id = %s
            ORDER BY pr.average_score DESC NULLS LAST
        """, (contest_id,))
    else:
        cursor.execute("""
            SELECT DISTINCT 
                a.id as application_id,
                a.project_name,
                a.full_name as applicant_name,
                a.direction,
                pr.average_score as final_score,
                pr.final_status,
                pr.report_data,
                pr.anomaly_detected
            FROM applications a
            LEFT JOIN project_ratings pr ON a.id = pr.application_id
            LEFT JOIN project_experts pe ON a.id = pe.application_id
            WHERE pe.id IS NOT NULL OR pr.id IS NOT NULL
            ORDER BY pr.average_score DESC NULLS LAST
        """)

    projects = cursor.fetchall()
    reports = []

    for project in projects:
        # Получаем оценки экспертов для этого проекта
        cursor.execute("""
            SELECT 
                es.*,
                e.full_name as expert_name
            FROM expert_scores es
            JOIN experts e ON es.expert_id = e.id
            JOIN project_experts pe ON es.project_expert_id = pe.id
            WHERE pe.application_id = %s
        """, (project['application_id'],))

        scores = cursor.fetchall()

        expert_scores = []
        anomalies_list = []

        # Группируем оценки по экспертам
        expert_data = {}
        for score in scores:
            if score['expert_name'] not in expert_data:
                expert_data[score['expert_name']] = {}
            expert_data[score['expert_name']][score['criteria_name']] = score['score']

        for expert_name, criteria_scores in expert_data.items():
            scores_list = list(criteria_scores.values())
            avg = round(sum(scores_list) / len(scores_list), 1) if scores_list else 0
            expert_scores.append({
                'name': expert_name,
                'scores': criteria_scores,
                'average': avg
            })

        # Проверяем аномалии
        if project['report_data']:
            try:
                report_data = json.loads(project['report_data']) if isinstance(project['report_data'], str) else project['report_data']
                if 'anomalies' in report_data:
                    for anomaly in report_data['anomalies']:
                        anomalies_list.append(
                            f"{anomaly.get('criteria', 'Критерий')}: эксперт {anomaly.get('expert_name', '')} поставил {anomaly.get('score', 0)} (среднее {anomaly.get('mean', 0)})")
            except:
                pass

        # Нормализуем final_score (если больше 100 - делим на 10)
        final_score_raw = project['final_score']
        if final_score_raw and final_score_raw > 100:
            final_score_normalized = round(final_score_raw / 10, 1)
        else:
            final_score_normalized = final_score_raw

        # Рекомендации на основе оценок
        recommendations = []
        if final_score_normalized and final_score_normalized >= 85:
            recommendations.append("Проект имеет высокий потенциал, рекомендован к финансированию в приоритетном порядке")
        elif final_score_normalized and final_score_normalized >= 70:
            recommendations.append("Проект рекомендован к финансированию при наличии бюджета")
        elif final_score_normalized and final_score_normalized >= 50:
            recommendations.append("Проект требует доработки перед финальным утверждением")
        else:
            recommendations.append("Проект не рекомендован к финансированию")

        reports.append({
            'project_name': project['project_name'],
            'applicant_name': project['applicant_name'],
            'direction': project['direction'],
            'final_score': final_score_normalized,
            'final_status': project['final_status'] or 'pending',
            'expert_scores': expert_scores,
            'anomalies': anomalies_list,
            'recommendations': recommendations
        })

    cursor.close()
    conn.close()

    # Статистика
    stats = {
        'total': len(reports),
        'approved': sum(1 for r in reports if r['final_status'] == 'approved' or r['final_status'] == 'Победитель' or r['final_status'] == 'Рекомендовано'),
        'rejected': sum(1 for r in reports if r['final_status'] == 'rejected' or r['final_status'] == 'Отклонено'),
        'avg_score': round(sum(r['final_score'] for r in reports if r['final_score']) / len(reports), 1) if reports else 0
    }

    return render_template("final_hearing_report.html",
                         reports=reports,
                         stats=stats,
                         report_date=datetime.now().strftime('%d.%m.%Y %H:%M'))

@app.route("/admin/export_final_report/<format>")
@admin_required
def export_final_report(format):
    """Экспорт итогового отчета после заслушивания"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем те же данные что и для отчета
    cursor.execute("""
        SELECT DISTINCT 
            a.id as application_id,
            a.project_name,
            a.full_name as applicant_name,
            a.direction,
            pr.average_score as final_score,
            pr.final_status,
            pr.report_data
        FROM applications a
        LEFT JOIN project_ratings pr ON a.id = pr.application_id
        LEFT JOIN project_experts pe ON a.id = pe.application_id
        WHERE pe.id IS NOT NULL OR pr.id IS NOT NULL
        ORDER BY pr.average_score DESC NULLS LAST
    """)

    projects = cursor.fetchall()

    if format == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Итоговый отчет после заслушивания"

        # Заголовки
        headers = ['№', 'Проект', 'Заявитель', 'Направление', 'Эксперт 1', 'Эксперт 1 - средний балл',
                   'Эксперт 2', 'Эксперт 2 - средний балл', 'Эксперт 3', 'Эксперт 3 - средний балл',
                   'Итоговый балл', 'Решение', 'Рекомендации']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 2
        for idx, project in enumerate(projects, 1):
            # Получаем экспертов и их средние баллы
            cursor.execute("""
                SELECT e.full_name, AVG(es.score) as avg_score
                FROM expert_scores es
                JOIN experts e ON es.expert_id = e.id
                JOIN project_experts pe ON es.project_expert_id = pe.id
                WHERE pe.application_id = %s
                GROUP BY e.id, e.full_name
            """, (project['application_id'],))

            experts = cursor.fetchall()

            # Формируем строку
            row_data = [
                idx,
                project['project_name'][:50],
                project['applicant_name'],
                project['direction']
            ]

            for i in range(3):
                if i < len(experts):
                    row_data.append(experts[i]['full_name'])
                    row_data.append(round(experts[i]['avg_score'], 1) if experts[i]['avg_score'] else 0)
                else:
                    row_data.append('—')
                    row_data.append('—')

            final_score = int(project['final_score'] * 10) if project['final_score'] else 0
            row_data.append(final_score)

            status_text = "Рекомендован" if project['final_status'] == 'approved' else (
                "Отклонен" if project['final_status'] == 'rejected' else "На доработке")
            row_data.append(status_text)

            recommendations = []
            if final_score >= 80:
                recommendations.append("Приоритетное финансирование")
            elif final_score >= 65:
                recommendations.append("Финансирование при наличии бюджета")
            elif final_score >= 50:
                recommendations.append("Требуется доработка")
            else:
                recommendations.append("Не рекомендован")
            row_data.append(", ".join(recommendations))

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)

            row += 1

        # Автоширина
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=final_hearing_report.xlsx'}
        )

    cursor.close()
    conn.close()
    return "Неверный формат", 400


@app.route("/admin/recheck_application/<int:app_id>")
@admin_required
def recheck_application(app_id):
    """Повторная AI проверка отдельной заявки"""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Получаем заявку
    cursor.execute("SELECT * FROM applications WHERE id = %s", (app_id,))
    app = cursor.fetchone()

    if not app:
        return "Заявка не найдена", 404

    # Выполняем AI проверку
    ai_result = ai_validate_application_detailed(dict(app))

    # Определяем новый статус
    new_status = 'На рассмотрении'
    if ai_result['can_approve'] and ai_result['total_score'] >= 75:
        new_status = 'Одобрено'
    elif ai_result['total_score'] < 50:
        new_status = 'Отклонено'

    # Обновляем заявку (обновляем formal_score)
    cursor.execute("""
        UPDATE applications 
        SET score = %s,
            formal_score = %s,
            validation_errors = %s,
            status = %s
        WHERE id = %s
    """, (
        ai_result['total_score'],
        ai_result['total_score'],
        json.dumps(ai_result, ensure_ascii=False, default=str),
        new_status,
        app_id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    # Перенаправляем обратно на страницу заявки
    return redirect(f"/admin/application/{app_id}")




# ================= ИНТЕГРАЦИЯ С МОДУЛЕМ ПОИСКА ПРОЕКТОВ =================

def _scouting_as_list(value):
    """Нормализует PostgreSQL ARRAY/list/string в обычный список строк."""
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(x).strip() for x in value if str(x).strip()]
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned.startswith('{') and cleaned.endswith('}'):
            cleaned = cleaned[1:-1]
        result = []
        current = []
        in_quotes = False
        for ch in cleaned:
            if ch == '"':
                in_quotes = not in_quotes
                continue
            if ch == ',' and not in_quotes:
                item = ''.join(current).strip().strip('"').strip("'")
                if item:
                    result.append(item)
                current = []
            else:
                current.append(ch)
        item = ''.join(current).strip().strip('"').strip("'")
        if item:
            result.append(item)
        return result
    text = str(value).strip()
    return [text] if text else []


def _serialize_contest_for_scouting(contest):
    """Преобразует строку contests в формат, который понимает модуль поиска."""
    contest_dict = dict(contest)
    directions = _scouting_as_list(contest_dict.get('directions'))
    name = contest_dict.get('name') or 'Без названия'
    description = contest_dict.get('description') or ''
    topic = (description + ' ' + ' '.join(directions)).strip()

    return {
        'id': contest_dict.get('id'),
        'title': name,
        'name': name,
        'description': description,
        'topic': topic,
        'goal': description,
        'directions': directions,
        'priority_topics': directions,
        'criteria': [
            'Актуальность',
            'Новизна',
            'Социальная или научно-техническая значимость',
            'Реализуемость',
            'Масштабируемость',
            'Обоснованность бюджета'
        ],
        'max_amount': contest_dict.get('max_amount'),
        'max_grant': contest_dict.get('max_amount'),
        'duration_months': contest_dict.get('duration_months'),
        'region': contest_dict.get('region'),
        'application_deadline': contest_dict.get('application_deadline').isoformat() if contest_dict.get('application_deadline') else None,
        'status': contest_dict.get('status'),
        'created_at': contest_dict.get('created_at').isoformat() if contest_dict.get('created_at') else None,
        'source_module': 'automation'
    }








@app.template_filter('from_json')
def from_json_filter(value):
    import json
    if value:
        try:
            return json.loads(value)
        except:
            return {}
    return {}


# === SCOUTING_CONTEST_ADMIN_START ===
SCOUTING_BASE_URL = os.getenv("SCOUTING_BASE_URL", "http://127.0.0.1:5000")
AUTOMATION_BASE_URL = os.getenv("AUTOMATION_BASE_URL", "http://127.0.0.1:8000")


def _parse_directions_for_scouting(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(x).strip() for x in value if str(x).strip()]
    raw = str(value).strip()
    if not raw:
        return []
    if raw.startswith("{") and raw.endswith("}"):
        raw = raw[1:-1]
    parts = []
    cur = []
    in_quotes = False
    for ch in raw:
        if ch == '"':
            in_quotes = not in_quotes
            continue
        if ch == "," and not in_quotes:
            item = "".join(cur).strip()
            if item:
                parts.append(item)
            cur = []
        else:
            cur.append(ch)
    item = "".join(cur).strip()
    if item:
        parts.append(item)
    return [p.strip().strip('"').strip("'") for p in parts if p.strip()]


def _contest_to_api_dict(contest):
    if not contest:
        return None
    item = dict(contest)
    directions = _parse_directions_for_scouting(item.get("directions"))
    description = item.get("description") or ""
    return {
        "id": item.get("id"),
        "title": item.get("name") or "",
        "name": item.get("name") or "",
        "description": description,
        "topic": " ".join([description, " ".join(directions)]).strip(),
        "goal": description,
        "directions": directions,
        "criteria": [
            "Актуальность",
            "Новизна",
            "Практическая значимость",
            "Реализуемость",
            "Соответствие направлениям конкурса",
            "Потенциал масштабирования"
        ],
        "priority_topics": directions,
        "max_amount": item.get("max_amount"),
        "max_grant": item.get("max_amount"),
        "duration_months": item.get("duration_months"),
        "region": item.get("region"),
        "application_deadline": item.get("application_deadline").isoformat() if item.get("application_deadline") else None,
        "status": item.get("status"),
    }


@app.route("/admin/go_to_search")
@admin_required
def go_to_search():
    contest_id = request.args.get("contest_id", type=int)

    # Если конкурс в админ-панели не выбран, открываем в модуле поиска страницу выбора конкурса.
    if not contest_id:
        return redirect(f"{SCOUTING_BASE_URL.rstrip('/')}/")

    return_url = f"{AUTOMATION_BASE_URL.rstrip('/')}/admin/dashboard?contest_id={contest_id}"
    search_url = (
        f"{SCOUTING_BASE_URL.rstrip('/')}/results"
        f"?external_competition_id={contest_id}"
        f"&return_url={return_url}"
    )
    return redirect(search_url)


@app.route("/admin/contests/create", methods=["GET", "POST"])
@admin_required
def admin_create_contest():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        directions_text = request.form.get("directions", "").strip()
        if not name or not description or not directions_text:
            return render_template("admin_contest_form.html", contest=None, directions_text=directions_text, error="Заполните название, описание и направления конкурса")
        directions = [x.strip() for x in directions_text.replace(";", "\n").splitlines() if x.strip()]
        max_amount = request.form.get("max_amount") or None
        duration_months = request.form.get("duration_months") or None
        region = request.form.get("region") or "Все регионы РФ"
        application_deadline = request.form.get("application_deadline") or None
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO contests (name, description, directions, max_amount, duration_months, region, application_deadline, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'active')
            RETURNING id
        """, (name, description, directions, max_amount, duration_months, region, application_deadline))
        contest_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(f"/admin/dashboard?contest_id={contest_id}")
    return render_template("admin_contest_form.html", contest=None, directions_text="")


@app.route("/admin/contests/<int:contest_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_edit_contest(contest_id):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM contests WHERE id = %s", (contest_id,))
    contest = cursor.fetchone()
    if not contest:
        cursor.close()
        conn.close()
        return redirect("/admin/dashboard")
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        directions_text = request.form.get("directions", "").strip()
        if not name or not description or not directions_text:
            cursor.close()
            conn.close()
            return render_template("admin_contest_form.html", contest=contest, directions_text=directions_text, error="Заполните название, описание и направления конкурса")
        directions = [x.strip() for x in directions_text.replace(";", "\n").splitlines() if x.strip()]
        cursor.execute("""
            UPDATE contests
            SET name=%s, description=%s, directions=%s, max_amount=%s, duration_months=%s, region=%s, application_deadline=%s
            WHERE id=%s
        """, (
            name,
            description,
            directions,
            request.form.get("max_amount") or None,
            request.form.get("duration_months") or None,
            request.form.get("region") or "Все регионы РФ",
            request.form.get("application_deadline") or None,
            contest_id
        ))
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(f"/admin/dashboard?contest_id={contest_id}")
    directions_text = "\n".join(_parse_directions_for_scouting(contest["directions"]))
    cursor.close()
    conn.close()
    return render_template("admin_contest_form.html", contest=contest, directions_text=directions_text)


@app.route("/admin/contests/<int:contest_id>/delete", methods=["POST"])
@admin_required
def admin_delete_contest(contest_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM contests WHERE id = %s", (contest_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect("/admin/dashboard")


@app.route("/api/scouting/health")
def api_scouting_health():
    return {"success": True, "service": "automation_module", "status": "online"}


@app.route("/api/scouting/contests")
def api_scouting_contests():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM contests WHERE COALESCE(status, 'active') = 'active' ORDER BY id")
    contests = [_contest_to_api_dict(row) for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return {"success": True, "contests": contests}


@app.route("/api/scouting/contests/<int:contest_id>")
def api_scouting_contest(contest_id):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM contests WHERE id = %s", (contest_id,))
    contest = cursor.fetchone()
    cursor.close()
    conn.close()
    if not contest:
        return {"success": False, "error": "Конкурс не найден"}, 404
    return {"success": True, "contest": _contest_to_api_dict(contest)}
# === SCOUTING_CONTEST_ADMIN_END ===


if __name__ == "__main__":
    port = int(os.environ.get('PORT', 8000))
    app.run(debug=False, host='0.0.0.0', port=port)


